## This Python script runs automated TRNSYS simulation
#   prepared by Hakan İbrahim Tol, PhD and Len Rijvers,  PDEng at TU/e on 22/02/2020

#  Libraries Imported
import subprocess           # to run the TRNSYS simulation
import shutil               # to duplicate the output txt file
import time                 # to measure the computation time
import pytest               # for testing
import pandas as pd
import numpy as np
from numpy.testing import assert_array_equal
from openpyxl import load_workbook
import pygfunction as gt
from scipy.optimize import minimize
from scipy.optimize import fsolve

def trnsys_results():
    deck_file_name = 'parallel_sync.dck'
    
    subprocess.run([r"C:\Trnsys18\Exe\TRNExe64.exe",r"C:\TRNSYS18\TRNLib\CallingPython-Cffi\Examples\08-ParallelSyncBoreholes\parallel_sync.dck","/h"])

def objective_function(x, T_in, m_flow, cp_f, T_g, LoadAgg, H_list):
    Rb = 0.08
    LoadAgg.set_current_load(x/sum(H_list))
    deltaT_b = LoadAgg.temporal_superposition()
    T_b = T_g - deltaT_b

    Tf = T_b - x/sum(H_list) * Rb
    T_f_in_single = Tf - ( x/2/m_flow/cp_f)
    
    return abs(T_f_in_single - T_in)

def python_results():
    # Load the borehole properties
    wb = load_workbook("GeoInput.xlsx")
    sheet = wb['Borehole']

    filled_rows = 0
    for row in sheet.iter_rows():
        if any(cell.value is not None for cell in row):
            filled_rows += 1
 
    H_list = [None] * (filled_rows - 1)
    D_list = [None] * (filled_rows - 1)
    rb_list = [None] * (filled_rows - 1)
    x_list = [None] * (filled_rows - 1)
    y_list = [None] * (filled_rows - 1)

    for row in range(2,filled_rows + 1):
        H_list[row-2] = sheet[f"A{row}"].value
        D_list[row-2] = sheet[f"B{row}"].value  
        rb_list[row-2] = sheet[f"C{row}"].value  
        x_list[row-2] = sheet[f"D{row}"].value
        y_list[row-2] = sheet[f"E{row}"].value   

    # Create borehole object
    borefield = [gt.boreholes.Borehole(H, D, rb, x, y) for H, D, rb, x, y in zip(H_list, D_list, rb_list, x_list, y_list)]

    # Load the ground properties
    sheet = wb['Ground']
    k = float(sheet['A2'].value)
    rho = float(sheet['B2'].value)
    cp = float(sheet['C2'].value)

    T_g = 8
    
    m_flow_borehole = 0.6     # Total fluid mass flow rate (kg/s)
    m_flow_network = m_flow_borehole * len(borefield)

    fluid = gt.media.Fluid('MPG', 20.)
    cp_f = fluid.cp     # Fluid specific isobaric heat capacity (J/kg.K)
    den_f = fluid.rho   # Fluid density (kg/m3)
    visc_f = fluid.mu   # Fluid dynamic viscosity (kg/m.s)
    k_f = fluid.k       # Fluid thermal conductivity (W/m.K)

    # Simulation parameters (must be consistent with TRNSYS!)
    dt = 3600.
    tmax = 200 * 3600.
    Nt = int(np.ceil(tmax/dt))
    time = dt * np.arange(1,Nt+1)

    LoadAgg = gt.load_aggregation.ClaessonJaved(dt,tmax)
    time_req = LoadAgg.get_times_for_simulation()

    gFunc = gt.gfunction.gFunction(borefield, k/rho/cp, time=time_req)
    LoadAgg.initialize(gFunc.gFunc/(2*np.pi*k))

    dT = np.zeros(Nt)
    Tf_in = np.zeros(Nt)
    Tf_out = np.zeros(Nt)
    Q_tot = np.zeros(Nt) + sum(H_list)*10
    Tin = -3
    for i in range(0,Nt):

        LoadAgg.next_time_step(i * 3600.)
        
        solution = minimize(objective_function, Q_tot[i], args = (Tin, m_flow_network, cp_f, T_g, LoadAgg, H_list))

        Q_tot[i] = solution.x[0]

        LoadAgg.set_current_load(Q_tot[i] /sum(H_list))
        deltaT_b = LoadAgg.temporal_superposition()
        T_b = T_g - deltaT_b

        Tf = T_b - Q_tot[i]/sum(H_list)*0.08
        Tf_in[i] = Tf - ( Q_tot[i]/2/m_flow_network/cp_f)

        Tf_out[i] = Tf + ( Q_tot[i]/2/m_flow_network/cp_f)


    np.savetxt('array_data.txt',Tf_out)

    return Tf_out

def test_parallel_sync():
    
    # Create results with trnsys
    trnsys_results()
    results_trnsys = pd.read_csv('parallel_sync.txt', delim_whitespace=True, skiprows=1, names=["TIME", "Tout"])
    T_trnsys = np.array(results_trnsys['Tout'])

    # Create results with pygfunction
    T_python = python_results()


    # Compare
    np.testing.assert_allclose(T_python[0:8760], T_trnsys[1:8761], atol=1e-3)  # Adjust `atol` for your use case


if __name__ == "__main__":
    pytest.main()
