# Python module for the TRNSYS Type calling Python using CFFI
# Data exchange with TRNSYS uses a dictionary, called TRNData in this file (it is the argument of all functions).
# Data for this module will be in a nested dictionary under the module name,
# i.e. if this file is calle "MyScript.py", the inputs will be in TRNData["MyScript"]["inputs"]
# for convenience the module name is saved in thisModule
#
# MKu, 2022-02-15

import os
import pygfunction as gt
import numpy as np
from openpyxl import load_workbook
from scipy.optimize import minimize
from scipy.optimize import fsolve

thisModule = os.path.splitext(os.path.basename(__file__))[0]

# Initialization: function called at TRNSYS initialization
# ---------------------------------------------------------------------------------------------------------------------
def Initialization(TRNData):
    global borefield
    global LoadAgg
    global SingleUTube
    global H_list
    global Q_tot
    global objective_function
    global m_flow_network
    global T_g
    global cp_f
    global Tf_in
    global Tf_out
    
    # Load the borehole properties
    wb = load_workbook("GeoInput.xlsx")
    sheet = wb['Borehole']

    filled_rows = 0
    for row in sheet.iter_rows():
        if any(cell.value is not None for cell in row):
            filled_rows += 1
 
    H_list = [None] * (filled_rows - 1)
    D_list = [None] * (filled_rows - 1)
    rb_list = [None] * (filled_rows - 1)
    x_list = [None] * (filled_rows - 1)
    y_list = [None] * (filled_rows - 1)

    for row in range(2,filled_rows + 1):
        H_list[row-2] = sheet[f"A{row}"].value
        D_list[row-2] = sheet[f"B{row}"].value  
        rb_list[row-2] = sheet[f"C{row}"].value  
        x_list[row-2] = sheet[f"D{row}"].value
        y_list[row-2] = sheet[f"E{row}"].value   

    # Create borehole object
    borefield = [gt.boreholes.Borehole(H, D, rb, x, y) for H, D, rb, x, y in zip(H_list, D_list, rb_list, x_list, y_list)]

    # Load the ground properties
    sheet = wb['Ground']
    k = float(sheet['A2'].value)
    rho = float(sheet['B2'].value)
    cp = float(sheet['C2'].value)

    T_g = 8

    # Fluid properties
    m_flow_borehole = 0.6     # Total fluid mass flow rate (kg/s)
    m_flow_network = m_flow_borehole * len(borefield)
    # The fluid is propylene-glycol (20 %) at 20 degC
    fluid = gt.media.Fluid('MPG', 20.)
    cp_f = fluid.cp     # Fluid specific isobaric heat capacity (J/kg.K)
    den_f = fluid.rho   # Fluid density (kg/m3)
    visc_f = fluid.mu   # Fluid dynamic viscosity (kg/m.s)
    k_f = fluid.k       # Fluid thermal conductivity (W/m.K)

    # Pipe dimensions
    rp_out = 0.02    # Pipe outer radius (m)
    rp_in = 0.018      # Pipe inner radius (m)
    D_s = 0.026       # Shank spacing (m)
    epsilon = 1.0e-6    # Pipe roughness (m)

    pos_single = [(-D_s, 0.), (D_s, 0.)]

    # Pipe properties
    k_p = 0.3           # Pipe thermal conductivity (W/m.K)

    # Grout properties
    k_g = 3.0           # Grout thermal conductivity (W/m.K). 3 should be a good value to simulate water

    # Pipe thermal resistance
    R_p = gt.pipes.conduction_thermal_resistance_circular_pipe(
        rp_in, rp_out, k_p)
    # Fluid to inner pipe wall thermal resistance (Single U-tube and double
    # U-tube in series)
    h_f = gt.pipes.convective_heat_transfer_coefficient_circular_pipe(
        m_flow_borehole, rp_in, visc_f, den_f, k_f, cp_f, epsilon)
 
    R_f_ser = 1.0/(h_f*2*np.pi*rp_in)


    nSteps = TRNData[thisModule]["total number of time steps"]

    # Simulation parameters (must be consistent with TRNSYS!)
    n_hours = 200 
    dt = 3600.
    tmax = n_hours * 3600.
    Nt = int(np.ceil(tmax/dt))
    time = dt * np.arange(1,Nt+1)

    LoadAgg = gt.load_aggregation.ClaessonJaved(dt,tmax)
    time_req = LoadAgg.get_times_for_simulation()

    gFunc = gt.gfunction.gFunction(borefield, k/(rho*cp), time=time_req)
    LoadAgg.initialize(gFunc.gFunc/(2*np.pi*k))

    # Tf_fake = np.linspace(3,-2, 8760)
    Q_tot = np.zeros(n_hours) + sum(H_list)*10
    Tf_in = np.zeros(n_hours)
    Tf_out = np.zeros(n_hours)

    def objective_function(x, T_in, m_flow_network, cp_f, T_g, LoadAgg, H_list):

        # x is the total load [W]
        Rb = 0.08
        LoadAgg.set_current_load(x/sum(H_list))
        deltaT_b = LoadAgg.temporal_superposition()
        T_b = T_g - deltaT_b

        # T_f_in_single = SingleUTube.get_inlet_temperature(
        #                     x, T_b, m_flow, cp_f)
        Tf = T_b - x/sum(H_list) * Rb
        T_f_in_single = Tf - ( x/2/m_flow_network/cp_f)
        return abs(T_f_in_single - T_in)

    return


# StartTime: function called at TRNSYS starting time (not an actual time step, initial values should be reported)
# ----------------------------------------------------------------------------------------------------------------------
def StartTime(TRNData):

    with open("Result.txt","w") as file:
        # file.write(str(Tf_out[stepNo])+"\n")
        pass

    return

# Iteration: function called at each TRNSYS iteration within a time step
# ----------------------------------------------------------------------------------------------------------------------
def Iteration(TRNData):

    # Define local short names for convenience (this is optional)
    Tin = TRNData[thisModule]["inputs"][0]
    m_flow_network = TRNData[thisModule]["inputs"][1]
    stepNo = TRNData[thisModule]["current time step number"]

    LoadAgg.next_time_step(stepNo * 3600.)
    
    solution = minimize(objective_function, Q_tot[stepNo - 2], args = (Tin, m_flow_network, cp_f, T_g, LoadAgg, H_list))
    Q_tot[stepNo-1] = solution.x[0]

    LoadAgg.set_current_load(Q_tot[stepNo-1] /sum(H_list))
    deltaT_b = LoadAgg.temporal_superposition()
    T_b = T_g - deltaT_b

    Tf = T_b - Q_tot[stepNo-1]/sum(H_list)*0.08
    Tf_in[stepNo -1] = Tf - ( Q_tot[stepNo-1]/2/m_flow_network/cp_f)

    Tf_out[stepNo -1] = Tf + ( Q_tot[stepNo-1]/2/m_flow_network/cp_f)


    # Calculate the outputs

    # Set outputs in TRNData
    TRNData[thisModule]["outputs"][0] = Tf_out[stepNo -1]
    TRNData[thisModule]["outputs"][1] = Q_tot[stepNo -1]

    with open("Result.txt","a") as file:
        file.write(str(Tf_out[stepNo -1] )+"\n")

    return

# EndOfTimeStep: function called at the end of each time step, after iteration and before moving on to next time step
# ----------------------------------------------------------------------------------------------------------------------
def EndOfTimeStep(TRNData):

    # This model has nothing to do during the end-of-step call
    
    return


# LastCallOfSimulation: function called at the end of the simulation (once) - outputs are meaningless at this call
# ----------------------------------------------------------------------------------------------------------------------
def LastCallOfSimulation(TRNData):

    # NOTE: TRNSYS performs this call AFTER the executable (the online plotter if there is one) is closed. 
    # Python errors in this function will be difficult (or impossible) to diagnose as they will produce no message.
    # A recommended alternative for "end of simulation" actions it to implement them in the EndOfTimeStep() part, 
    # within a condition that the last time step has been reached.
    #
    # Example (to be placed in EndOfTimeStep()):
    #
    # stepNo = TRNData[thisModule]["current time step number"]
    # nSteps = TRNData[thisModule]["total number of time steps"]
    # if stepNo == nSteps-1:     # Remember: TRNSYS steps go from 0 to (number of steps - 1)
    #     do stuff that needs to be done only at the end of simulation

    return